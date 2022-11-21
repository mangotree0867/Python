import random
from datetime import date
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import openpyxl

listofnumbers = [1,2,3,4,5,6,7,8,9,0]                   #list for 10% reward lottery
rewards = 0                                             #number of rewards
completed = ""                                          #if completed or not
count = 0                                               #number of habits done today

#getting today's date
today = date.today()
record = today.strftime("%m-%d-%y")

#opening file
file_name = "Habits.xlsx"
sheet = file_name
wb = openpyxl.load_workbook(file_name)
ws = wb.active

col_write = ws.cell(row = 24, column = 4).value + 5     #getting the number of days passed
# writing today's date
date_cell = ws.cell(row = 1, column= col_write)        
date_cell.value = record

#iterate until next cell is blank
for i in range(4, ws.max_row+1):

    #get habits in order
    cell_obj = ws.cell(row=i, column=1)
    print("Did you " + cell_obj.value , end="?  yes or no :\n")
    
    #get user input
    completed = input()

    #act depending on the user input
    if (completed == "yes") or (completed == "Yes") or (completed =="y") or (completed =="Y"):
        count+=1
        #write in yes in the excel file
        ws.cell(row= i, column = col_write).value = "yes"
        
        #pick a number from 0~9
        number = random.choice(listofnumbers)
        print(number)
        #if number is 0 you get rewarded
        if(number == 0):
            rewards+=1
    else:
        ws.cell(row=i, column=col_write).value = "no"
    
    #break if the next habit cell is empty
    cell_obj_next = ws.cell(row=i + 1, column=1)
    if(cell_obj_next.value == None):
        break

#get the number of 'Days passed'
Days = ws.cell(row = 24, column = 4)
#increases number every time this runs
Days.value += 1

#write count for Habits made today
ws.cell(row = 20, column=col_write).value = count
#write rewards for Rewards earned today
ws.cell(row = 22, column=col_write).value = rewards

reward_type = ""
#choose reward type if you have rewards
for i in range(0, rewards):
    reward_type = input("Choose your reward : Money, Games, Youtube")

    if(reward_type == "Money") or (reward_type =="money") or (reward_type =="m"):
        ws.cell(row = 27, column=3).value +=1
    elif(reward_type == "Games") or (reward_type =="games") or (reward_type =="g"):
        ws.cell(row = 28, column= 3).value +=1
    elif(reward_type == "Youtube") or (reward_type =="youtube") or (reward_type =="y"):
        ws.cell(row=29, column=3).value += 1
    else:
        print("wrong input please try again")

wb.save(file_name)